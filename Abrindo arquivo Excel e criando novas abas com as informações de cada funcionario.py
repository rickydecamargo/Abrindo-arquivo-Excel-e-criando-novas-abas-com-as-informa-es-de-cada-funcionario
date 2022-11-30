#Criando vários Arquivos Excel para cada um dos vendedores
#Deixe a pasta apenas com o arquivo Quebrar e rode para ver outros 4 arquivos sendo gerados ao rodar o script.
#Neste exemplo é necessário que o arquivo excel já esta preenchido na primeira página
#https://www.udemy.com/course/python-rpa-e-excel-aprenda-automatizar-processos-e-planilhas/learn/lecture/27889072#overview

from openpyxl import load_workbook
from openpyxl import Workbook
import os

nome_arquivo = "C:\\Users\\Windows\\Desktop\\Python Projetos\\openpyxl\\ExcelEmail\Quebrar.xlsx"
planilha_aberta = load_workbook(filename=nome_arquivo)

#Seleciona a Sheet de Dados
sheet_selecionada = planilha_aberta['Dados']

criandoNovoArquivoExcel = Workbook()

nomeNovo = ""
totalLinha = len(sheet_selecionada['A']) + 1

#Para tornar dinâmico a leitura das linhas, para ler todas que estiverem com informações
for linha in range(2, len(sheet_selecionada['A']) + 1):

    nomeAtual = sheet_selecionada['A%s' % linha].value

    if nomeNovo == nomeAtual:

        linhaSheetQuebra = len(selecionaSheetVendasNovaPlanilha['A']) + 1
        celulaColunaA = "A" + str(linhaSheetQuebra)
        celulaColunaB = "B" + str(linhaSheetQuebra)
        celulaColunaC = "C" + str(linhaSheetQuebra)

        #Preenche os dados
        selecionaSheetVendasNovaPlanilha[celulaColunaA] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaB] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha[celulaColunaC] = sheet_selecionada['C%s' % linha].value

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)

    else:
        #Adiciona o nome do funcionario que esta na linha que o código está passando
        nomeNovo = sheet_selecionada['A%s' % linha].value

        nova_planilha = criandoNovoArquivoExcel.active

        nova_planilha.title = "Vendas"

        caminhoNovaPlanilha = "C:\\Users\\Windows\\Desktop\\Python Projetos\\openpyxl\\ExcelEmail\\" + sheet_selecionada['A%s' % linha].value +".xlsx"

        selecionaSheetVendasNovaPlanilha = criandoNovoArquivoExcel['Vendas']

        #Coloca os titulos
        selecionaSheetVendasNovaPlanilha['A1'] = "Vendedor"
        selecionaSheetVendasNovaPlanilha['B1'] = "Produtos"
        selecionaSheetVendasNovaPlanilha['C1'] = "Vendas"

        #Preenche as informações na segunda linha
        selecionaSheetVendasNovaPlanilha['A2'] = sheet_selecionada['A%s' % linha].value
        selecionaSheetVendasNovaPlanilha['B2'] = sheet_selecionada['B%s' % linha].value
        selecionaSheetVendasNovaPlanilha['C2'] = sheet_selecionada['C%s' % linha].value

        selecionaSheetVendasNovaPlanilha.delete_rows(3,100000)

        criandoNovoArquivoExcel.save(filename=caminhoNovaPlanilha)